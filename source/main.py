from pathlib import Path
from copy import deepcopy
from worker import Worker

import argparse
import os
import json
import openpyxl


class ArgumentList:
    _data = {
        'ss': ([], {'prog': 'Data Analyser', 'description': 'Fills in Excel column data based on established rules'}),
        '-l': (['load_path'], {'help': ''}),
        '-s': (['save_path'], {'help': ''}),
        '-r': (['-r', '--rule'], {'default': 'settings/rule/data.json', 'help': ''}),
        '-d': (['-v', '--validation'], {'nargs': '?', 'const': 'settings/validation/data.json', 'help': ''}),
        '-c': (['--create'], {'nargs': '1', 'help': ''})
    }

    @staticmethod
    def get_args(item):
        result = ArgumentList._data.get(item)
        return deepcopy(result[0]) if result is not None else None

    @staticmethod
    def get_kwargs(item):
        result = ArgumentList._data.get(item)
        return deepcopy(result[1]) if result is not None else None


def main(load_path, save_path, rule_path, validation_path):
    if not os.path.exists(load_path):
        raise AttributeError(f'File or directory "{load_path}" does not exist!')
    if validation_path is not None and not os.path.exists(validation_path):
        raise AttributeError(f'File or directory "{validation_path}" does not exist!')
    if not os.path.exists(rule_path):
        raise RuntimeError(f'Error catch while trying to get a rule at "{rule_path}"!')

    with open(rule_path, 'r', encoding='utf-8') as f:
        settings = json.load(f)

    book = openpyxl.load_workbook(settings['book'])
    data = {i['source_sheet']: [] for i in settings['rules']}
    for i in settings['rules']:
        data[i['source_sheet']] += [Worker(book[i['rule_sheet']], i['source_column'], i['rule_column'], i['results'])]
    book.close()

    if os.path.isdir(load_path):
        for i in os.listdir(load_path):
            if i.split('.')[-1] == 'xlsx':
                book = openpyxl.load_workbook(Path(load_path) / Path(i))
                for k in data.keys():
                    for n in data[k]:
                        n.process_sheet(book[k])
                book.save(Path(load_path) / Path(i))
                book.close()
    elif os.path.isfile(load_path) and load_path.split('.')[-1] == 'xlsx':
        book = openpyxl.load_workbook(load_path)
        for k in data.keys():
            for n in data[k]:
                n.process_sheet(book[k])
        book.save(save_path)
        book.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(*ArgumentList.get_kwargs('ss'))
    parser.add_argument(*ArgumentList.get_args('-l'), *ArgumentList.get_kwargs('-l'))
    parser.add_argument(*ArgumentList.get_args('-s'), *ArgumentList.get_kwargs('-s'))
    parser.add_argument(*ArgumentList.get_args('-r'), *ArgumentList.get_kwargs('-r'))
    parser.add_argument(*ArgumentList.get_args('-d'), *ArgumentList.get_kwargs('-d'))
    args = parser.parse_args()

    main(args.load_path, args.save_path, args.rule_path, args.validation_path)
