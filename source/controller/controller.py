from source.model.worker import Worker
from source.environment import *

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

import json
import pathlib


class Controller:

    def __init__(self, load_path, save_path=None):
        if not os.path.exists(load_path := pathlib.Path(load_path)):
            raise AttributeError(f'Load path "{load_path}" does not exist!')
        if load_path.is_file():
            if load_path.name.split('.')[-1] != 'xlsx':
                raise AttributeError(f'Load file "{load_path}" must have ".xlsx" extend!')
            save_path = pathlib.Path(DEFAULT_SAVE_PATH if save_path is None else save_path)
            if save_path.name.split('.')[-1] != 'xlsx':
                raise AttributeError(f'Save file "{save_path}" must have ".xlsx" extend!')

        self.load_path = load_path
        self.save_path = save_path

    def column_fill(self, rule_path):
        if not os.path.exists(rule_path := pathlib.Path(rule_path)):
            raise AttributeError(f'(CF) Rule path "{rule_path}" does not exist!')

        with open(rule_path, 'r', encoding='utf-8') as f:
            settings = json.load(f)

        book = openpyxl.load_workbook(settings['book'])
        data = {i['source_sheet']: [] for i in settings['rules']}
        for i in settings['rules']:
            data[i['source_sheet']] += [Worker(book[i['rule_sheet']], i['source_column'], i['rule_column'],
                                               i['results'], i['rule_size'])]
        book.close()

        if os.path.isdir(self.load_path):
            for i in os.listdir(self.load_path):
                if i.split('.')[-1] == 'xlsx':
                    book = openpyxl.load_workbook(self.load_path / pathlib.Path(i))
                    for k in data.keys():
                        for n in data[k]:
                            n.process_sheet(book[k])
                    book.save(self.save_path / pathlib.Path(i))
                    book.close()
        else:
            book = openpyxl.load_workbook(self.load_path)
            for k in data.keys():
                for n in data[k]:
                    n.process_sheet(book[k])
            book.save(self.save_path)
            book.close()

    def validation_add(self, rule_path):
        if not os.path.exists(rule_path := pathlib.Path(rule_path)):
            raise AttributeError(f'(VA) Rule path "{rule_path}" does not exist!')

        with open(rule_path, 'r', encoding='utf-8') as f:
            settings = json.load(f)

        if os.path.isdir(self.load_path):
            for i in os.listdir(self.load_path):
                if i.split('.')[-1] == 'xlsx':
                    book = openpyxl.load_workbook(self.load_path / pathlib.Path(i))
                    for k in settings:
                        rule = DataValidation(type="list", formula1=f'=\'{k["sheet_from"]}\'!{k["from"]}',
                                              allow_blank=True)
                        book[k['sheet_to']].add_data_validation(rule)
                        rule.add(k['to'])
                    book.save(self.save_path / pathlib.Path(i))
                    book.close()
        else:
            book = openpyxl.load_workbook(self.load_path)
            for k in settings:
                rule = DataValidation(type="list", formula1=f'=\'{k["sheet_from"]}\'!{k["from"]}', allow_blank=True)
                book[k['sheet_to']].add_data_validation(rule)
                rule.add(k['to'])
            book.save(self.save_path)
            book.close()
